"""
check_history.py - 读取运行日志，并按 1~7 天窗口选取对比基准
"""
import argparse
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

from run_context import parse_run_path, result_dir_from_run_path, datetime_display


def _split_paths(raw: str) -> list:
    return [p.strip() for p in str(raw or "").replace("；", ";").split(";") if p.strip()]


def _run_datetime(info: dict) -> datetime:
    date = info.get("date", "")
    time = (info.get("time") or "0000").zfill(4)
    if not date or len(date) != 8:
        return datetime.min
    try:
        return datetime.strptime(f"{date}{time}", "%Y%m%d%H%M")
    except ValueError:
        return datetime.strptime(date, "%Y%m%d")


def _load_runs_from_log(ws, asin: str, output_dir: str) -> list:
    runs = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] != asin:
            continue
        run_time = str(row[1] or "")
        data_date = str(row[2] or "")
        path = _split_paths(str(row[3] or ""))[0] if row[3] else ""
        if not path:
            continue
        info = parse_run_path(path)
        if not info.get("date"):
            info["date"] = data_date
        info["path"] = path
        info["run_time"] = run_time
        info["datetime_display"] = run_time or info.get("datetime_display", "")
        info["result_dir"] = result_dir_from_run_path(output_dir, path)
        info["_dt"] = _run_datetime(info)
        runs.append(info)
    runs.sort(key=lambda r: r["_dt"])
    return runs


def _load_runs_legacy(ws, asin: str, output_dir: str) -> list:
    headers = [cell.value for cell in ws[1]]
    asin_col = next((i for i, h in enumerate(headers) if h and "ASIN" in str(h).upper()), None)
    if asin_col is None:
        return []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[asin_col] != asin:
            continue
        last_snap = str(row[4]) if len(row) > 4 and row[4] else ""
        hist_raw = str(row[5]) if len(row) > 5 and row[5] else last_snap
        path_list = _split_paths(hist_raw)
        if last_snap and last_snap not in path_list:
            path_list.append(last_snap)
        runs = []
        for p in path_list:
            info = parse_run_path(p)
            info["path"] = p
            info["result_dir"] = result_dir_from_run_path(output_dir, p)
            info["run_time"] = info.get("datetime_display", "")
            info["_dt"] = _run_datetime(info)
            runs.append(info)
        runs.sort(key=lambda r: r["_dt"])
        return runs
    return []


def _parse_data_date(date_str: str):
    if not date_str or len(date_str) != 8:
        return None
    try:
        return datetime.strptime(date_str, "%Y%m%d").date()
    except ValueError:
        return None


def _fill_compare_out(out: dict, prev: dict, days: int, in_window: bool, note: str) -> dict:
    out.update({
        "compare_prev_run_path": prev["path"],
        "compare_prev_date": prev.get("date", ""),
        "compare_prev_datetime": prev.get("run_time") or prev.get("datetime_display", ""),
        "compare_prev_result_dir": prev.get("result_dir", ""),
        "compare_days_apart": days,
        "compare_in_window": in_window,
        "compare_note": note,
    })
    return out


def find_compare_baseline(runs: list, curr_date: str = "", min_days: int = 1, max_days: int = 7) -> dict:
    """
    词库对比按「原数据日期」选基准，不是按同日重复跑库时间。

    优先级：
    1. 与本次数据日期不同、且在 1~7 天窗口内的最近一次数据日（如 0610 → 0611）
    2. 仅当没有可比的异日数据时，才对比同日前一次运行（同日重跑）
    3. 回退到紧邻上一次运行（可能超出窗口，报告会标注 warning）
    """
    out = {
        "compare_prev_run_path": "",
        "compare_prev_date": "",
        "compare_prev_datetime": "",
        "compare_prev_result_dir": "",
        "compare_days_apart": None,
        "compare_in_window": False,
        "compare_note": "",
    }
    if len(runs) < 2:
        out["compare_note"] = "仅一次运行，无法对比"
        return out

    if curr_date:
        curr_candidates = [r for r in runs if r.get("date") == curr_date]
        curr = curr_candidates[-1] if curr_candidates else runs[-1]
    else:
        curr = runs[-1]

    curr_dt = curr["_dt"]
    curr_data_date = curr.get("date", "")
    curr_dd = _parse_data_date(curr_data_date)
    priors = [r for r in runs if r["_dt"] < curr_dt]
    if not priors:
        out["compare_note"] = "当前为最早一次运行"
        return out

    # 1) 异日数据：按原数据日期差在 1~7 天窗口内选最近一期
    if curr_dd:
        best_by_date: dict[str, dict] = {}
        for r in priors:
            prev_date = r.get("date", "")
            if not prev_date or prev_date == curr_data_date:
                continue
            prev_dd = _parse_data_date(prev_date)
            if not prev_dd:
                continue
            days = (curr_dd - prev_dd).days
            if min_days <= days <= max_days:
                slot = best_by_date.get(prev_date)
                if slot is None or r["_dt"] > slot["run"]["_dt"]:
                    best_by_date[prev_date] = {"days": days, "run": r}
        if best_by_date:
            prev_date = max(best_by_date.keys())
            slot = best_by_date[prev_date]
            prev = slot["run"]
            days = slot["days"]
            return _fill_compare_out(
                out, prev, days, True,
                f"对比 {days} 天前数据（{prev_date} → {curr_data_date}，原数据日期）",
            )

    # 2) 同日前一次（仅异日不可比时：同日重跑同一批数据）
    same_day = [r for r in priors if r.get("date") == curr_data_date]
    if same_day:
        prev = same_day[-1]
        return _fill_compare_out(
            out, prev, 0, True,
            f"对比同日前一次运行（数据日期均为 {curr_data_date}）",
        )

    # 3) 回退：紧邻上一次（按数据日期计间隔）
    prev = priors[-1]
    prev_dd = _parse_data_date(prev.get("date", ""))
    if curr_dd and prev_dd:
        days = (curr_dd - prev_dd).days
    else:
        days = (curr_dt.date() - prev["_dt"].date()).days
    in_window = min_days <= days <= max_days if days is not None else False
    note = (
        f"对比 {days} 天前数据（{prev.get('date', '')} → {curr_data_date}，超出 {max_days} 天窗口）"
        if not in_window
        else f"对比 {days} 天前数据（{prev.get('date', '')} → {curr_data_date}）"
    )
    return _fill_compare_out(out, prev, days, in_window, note)


def check_history(asin: str, output_dir: str, curr_date: str = "", compare_window: int = 7) -> dict:
    history_file = os.path.join(output_dir, "ASIN历史记录.xlsx")
    result = {
        "has_history": False,
        "run_count": 0,
        "last_date": "",
        "last_time": "",
        "last_datetime": "",
        "last_run_path": "",
        "last_result_dir": "",
        "prev_run_path": "",
        "prev_path": "",
        "prev_datetime": "",
        "all_runs": [],
        "compare_window_days": compare_window,
    }

    if not os.path.exists(history_file) or not HAS_OPENPYXL:
        return result

    try:
        wb = openpyxl.load_workbook(history_file)
        if "运行日志" in wb.sheetnames:
            runs = _load_runs_from_log(wb["运行日志"], asin, output_dir)
        else:
            runs = _load_runs_legacy(wb.active, asin, output_dir)

        if not runs:
            wb.close()
            return result

        latest = runs[-1]
        prev = runs[-2] if len(runs) >= 2 else None
        compare = find_compare_baseline(runs, curr_date=curr_date, max_days=compare_window)

        result["has_history"] = len(runs) >= 2 or bool(prev)
        result["run_count"] = len(runs)
        result["last_date"] = latest.get("date", "")
        result["last_time"] = latest.get("time", "")
        result["last_datetime"] = latest.get("run_time") or latest.get("datetime_display", "")
        result["last_run_path"] = latest.get("path", "")
        result["last_result_dir"] = latest.get("result_dir", "")
        result["all_runs"] = [{k: v for k, v in r.items() if k != "_dt"} for r in runs]

        if prev:
            result["prev_run_path"] = prev.get("path", "")
            result["prev_path"] = prev.get("result_dir", "")
            result["prev_datetime"] = prev.get("run_time") or prev.get("datetime_display", "")

        result.update(compare)
        wb.close()
    except Exception as e:
        result["error"] = str(e)

    return result


def pd_excel_path(result_dir: str, asin: str, date: str) -> str:
    return os.path.join(result_dir, f"{asin}_PD主攻词单_{date}.xlsx")


def main():
    parser = argparse.ArgumentParser(description="检查 ASIN 历史记录")
    parser.add_argument("--asin", required=True)
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--curr-date", default="", help="本次数据日期 YYYYMMDD")
    parser.add_argument("--compare-window", type=int, default=7)
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    result = check_history(args.asin, args.output_dir, curr_date=args.curr_date,
                           compare_window=args.compare_window)
    print(json.dumps(result, ensure_ascii=False, indent=2))

    if result.get("all_runs"):
        print(
            f"\n[历史] {args.asin} 共 {result['run_count']} 次 | "
            f"最近 {result['last_datetime']} | {result['last_run_path']}",
            file=sys.stderr,
        )
        if result.get("compare_prev_run_path"):
            print(
                f"[对比基准] {result.get('compare_note')} | "
                f"{result.get('compare_prev_datetime')} | {result.get('compare_prev_run_path')}",
                file=sys.stderr,
            )


if __name__ == "__main__":
    main()
