"""
update_history.py - 更新 ASIN 历史记录表（每次运行追加一行，不覆盖）

Sheet「运行日志」：一行 = 一次跑库（同日期可多次，用 HHmm 区分）
"""
import argparse
import os
import re
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("请安装 openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

from run_context import datetime_display, parse_run_path, history_run_path

LOG_HEADERS = ["ASIN", "运行时间", "数据日期", "快照路径"]
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(color="FFFFFF", bold=True, name="Microsoft YaHei")
ALT_FILL = PatternFill(start_color="DEEAF1", end_color="DEEAF1", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def _normalize_datetime_cell(val) -> str:
    if val is None or str(val).strip() == "":
        return ""
    s = str(val).strip()
    if re.fullmatch(r"\d{8}", s):
        return datetime_display(s)
    if re.fullmatch(r"\d{8}/\d{4}", s):
        date, time = s.split("/")
        return datetime_display(date, time)
    return s


def _normalize_path(path: str) -> str:
    return path.strip().replace("\\", "/") if path else ""


def _resolve_history_path(run_path: str, asin: str, date: str, time: str) -> str:
    p = _normalize_path(run_path)
    info = parse_run_path(p)
    if info.get("time"):
        return p
    if info.get("asin") and info.get("date"):
        return history_run_path(info["asin"], info["date"], time)
    return history_run_path(asin, date, time)


def _ensure_log_sheet(wb):
    if "运行日志" in wb.sheetnames:
        ws = wb["运行日志"]
    else:
        ws = wb.create_sheet("运行日志", 0)
        for col_idx, h in enumerate(LOG_HEADERS, 1):
            cell = ws.cell(1, col_idx, value=h)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = THIN_BORDER
        ws.row_dimensions[1].height = 22
    return ws


def _existing_paths(ws) -> set[str]:
    paths = set()
    for row_idx in range(2, ws.max_row + 1):
        p = ws.cell(row_idx, 4).value
        if p:
            paths.add(_normalize_path(str(p)))
    return paths


def _migrate_legacy_sheet(wb, ws_log):
    """把旧版「汇总行+分号路径」迁移到运行日志"""
    if ws_log.max_row > 1:
        return
    legacy = wb.active if wb.active.title != "运行日志" else None
    if legacy is None:
        for name in wb.sheetnames:
            if name != "运行日志":
                legacy = wb[name]
                break
    if legacy is None or legacy.max_row < 2:
        return

    headers = [legacy.cell(1, c).value for c in range(1, legacy.max_column + 1)]
    if not headers or "ASIN" not in str(headers[0] or ""):
        return

    for row_idx in range(2, legacy.max_row + 1):
        asin = legacy.cell(row_idx, 1).value
        if not asin:
            continue
        first = _normalize_datetime_cell(legacy.cell(row_idx, 2).value)
        hist_raw = legacy.cell(row_idx, 6).value or legacy.cell(row_idx, 5).value or ""
        paths = [_normalize_path(p) for p in str(hist_raw).replace("；", ";").split(";") if p.strip()]
        snap = _normalize_path(str(legacy.cell(row_idx, 5).value or ""))
        if snap and snap not in paths:
            paths.append(snap)
        if not paths and snap:
            paths = [snap]

        for i, path in enumerate(paths):
            info = parse_run_path(path)
            run_time = first if i == 0 and first else datetime_display(
                info.get("date", ""), info.get("time") or None
            )
            if not info.get("date"):
                continue
            if path in _existing_paths(ws_log):
                continue
            row = ws_log.max_row + 1
            ws_log.cell(row, 1, value=str(asin))
            ws_log.cell(row, 2, value=run_time)
            ws_log.cell(row, 3, value=info["date"])
            ws_log.cell(row, 4, value=path)


def _style_log_row(ws, row_idx: int):
    fill = ALT_FILL if row_idx % 2 == 0 else None
    for col_idx in range(1, len(LOG_HEADERS) + 1):
        cell = ws.cell(row_idx, col_idx)
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical="center", wrap_text=(col_idx >= 4))
        cell.font = Font(name="Microsoft YaHei", size=10)
        if fill:
            cell.fill = fill


def update_history(asin: str, output_dir: str, date: str, time: str, run_path: str):
    history_file = os.path.join(output_dir, "ASIN历史记录.xlsx")
    dt_display = datetime_display(date, time)
    hist_path = _resolve_history_path(run_path, asin, date, time)

    if os.path.exists(history_file):
        wb = openpyxl.load_workbook(history_file)
    else:
        wb = openpyxl.Workbook()
        if "Sheet" in wb.sheetnames:
            del wb["Sheet"]

    ws = _ensure_log_sheet(wb)
    _migrate_legacy_sheet(wb, ws)

    if hist_path in _existing_paths(ws):
        print(f"ℹ️  快照路径已存在，跳过重复写入：{hist_path}")
    else:
        row = ws.max_row + 1
        ws.cell(row, 1, value=asin)
        ws.cell(row, 2, value=dt_display)
        ws.cell(row, 3, value=date)
        ws.cell(row, 4, value=hist_path)
        _style_log_row(ws, row)

    for col_idx, width in enumerate([18, 20, 14, 36], 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(history_file)
    run_count = max(0, ws.max_row - 1)
    print(f"OK history updated: {history_file}")
    print(f"   ASIN: {asin} | time: {dt_display} | path: {hist_path} | total runs: {run_count}")


def main():
    parser = argparse.ArgumentParser(description="更新 ASIN 历史记录（每次运行追加一行）")
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--migrate-only", action="store_true")
    parser.add_argument("--asin", default="")
    parser.add_argument("--date", default="", help="YYYYMMDD")
    parser.add_argument("--time", default="", help="HHmm")
    parser.add_argument("--run-path", default="", help="如 B08PNQCKF7/20260611")
    args = parser.parse_args()

    os.makedirs(args.output_dir, exist_ok=True)
    if args.migrate_only:
        history_file = os.path.join(args.output_dir, "ASIN历史记录.xlsx")
        if not os.path.exists(history_file):
            print(f"ERROR: not found {history_file}", file=sys.stderr)
            sys.exit(1)
        wb = openpyxl.load_workbook(history_file)
        ws = _ensure_log_sheet(wb)
        _migrate_legacy_sheet(wb, ws)
        for row_idx in range(2, ws.max_row + 1):
            _style_log_row(ws, row_idx)
        wb.save(history_file)
        print(f"OK migrated: {history_file} ({ws.max_row - 1} rows)")
        return

    if not args.asin or not args.date or not args.run_path:
        print("ERROR: need --asin --date --run-path", file=sys.stderr)
        sys.exit(1)
    time = args.time or datetime.now().strftime("%H%M")
    update_history(args.asin, args.output_dir, args.date, time, args.run_path)


if __name__ == "__main__":
    main()
