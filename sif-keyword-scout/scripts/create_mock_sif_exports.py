"""
create_mock_sif_exports.py  [开发/测试专用，不在主流程中调用]
生成符合 SOP 结构的测试用三张原始表（无真实 Sif 文件时用于链路验证）

用法：
  python create_mock_sif_exports.py --output-dir "其他/test/_mock_raw" --date 20260611
"""
import stdio_utf8  # noqa: F401

import argparse
import os
import pandas as pd
from openpyxl import Workbook


def write_table1(path: str):
    wb = Workbook()
    for sheet in ["高相关", "中相关", "低相关"]:
        ws = wb.create_sheet(sheet)
        ws.append([])  # 第1行空行
        ws.append(["关键词", "中文翻译", "相关性", "相关性得分", "周搜索量", "ABA排名",
                   "24小时转化率", "", "建议竞价中值", "", "Top3集中度"])
        if sheet == "高相关":
            rows = [
                ["stanley cup", "斯坦利杯", "高", 95, 360806, 1200, 0.12, "", 1.25, "", 0.22],
                ["stanley tumbler", "斯坦利保温杯", "高", 90, 11914, 3500, 0.08, "", 0.95, "", 0.28],
                ["tumbler with handle", "带把手保温杯", "高", 85, 8500, 8000, 0.06, "", 0.88, "", 0.31],
                ["40 oz tumbler", "40盎司杯", "高", 82, 6200, 12000, 0.05, "", 0.75, "", 0.27],
                ["insulated tumbler", "保温杯", "高", 78, 4100, 18000, 0.04, "", 0.70, "", 0.33],
                ["coffee tumbler", "咖啡杯", "高", 75, 2800, 25000, 0.03, "", 0.65, "", 0.35],
            ]
        elif sheet == "中相关":
            rows = [
                ["water bottle", "水瓶", "中", 55, 1500, 45000, 0.02, "", 0.55, "", 0.38],
                ["travel mug", "旅行杯", "中", 50, 900, 60000, 0.01, "", 0.50, "", 0.42],
            ]
        else:
            rows = [
                ["coach cover", "教练罩", "低", 10, 120, 500000, "", "", 0.30, "", 0.55],
            ]
        for r in rows:
            ws.append(r)
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    wb.save(path)


def write_table2(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "反查流量词"
    ws.append([])
    headers = ["", "关键词", "", "关键词效果类型", "全部流量占比", "", "", "自然流量占比",
               "", "", "", "广告流量占比"]
    # 补齐到66列
    while len(headers) < 66:
        headers.append("")
    headers[43] = "周ABA排名"
    headers[65] = "ABA TOP3集中度"
    ws.append(headers)
    rows = [
        ["", "stanley cup", "", "[搜索量同比增长]", 0.17, "", "", 0.08, "", "", "", 0.09, *[""]*31, 2561791, *[""]*21, 0.32],
        ["", "stanley tumbler", "", "[搜索量同比增长]", 0.12, "", "", 0.05, "", "", "", 0.07, *[""]*31, 3200000, *[""]*21, 0.28],
        ["", "tumbler with handle", "", "[搜索量同比增长]", 0.09, "", "", 0.03, "", "", "", 0.06, *[""]*31, 4100000, *[""]*21, 0.35],
        ["", "40 oz tumbler", "", "", 0.06, "", "", 0.15, "", "", "", 0.04, *[""]*31, 5000000, *[""]*21, 0.30],
        ["", "insulated tumbler", "", "", 0.04, "", "", 0.12, "", "", "", 0.03, *[""]*31, 6000000, *[""]*21, 0.38],
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)


def write_table3(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "查广告词"
    ws.append([])
    ws.append(["", "广告搜索词", "翻译", "SP广告流量贡献占比", "竞品在该词的SP广告流量份额",
               "", "", "", "", "", "搜索量"])
    rows = [
        ["", "stanley cup", "斯坦利杯", "37.27%", "3.21%", "", "", "", "", "", 360806],
        ["", "stanley tumbler", "斯坦利保温杯", "22.10%", "4.50%", "", "", "", "", "", 11914],
        ["", "tumbler with handle", "带把手保温杯", "8.50%", "2.80%", "", "", "", "", "", 8500],
        ["", "40 oz tumbler", "40盎司杯", "15.00%", "18.00%", "", "", "", "", "", 6200],
        ["", "water bottle", "水瓶", "5.00%", "1.20%", "", "", "", "", "", 1500],
        ["", "coach cover", "教练罩", "1.00%", "0.50%", "", "", "", "", "", 120],
    ]
    for r in rows:
        ws.append(r)
    wb.save(path)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", required=True)
    parser.add_argument("--date", required=True)
    args = parser.parse_args()
    os.makedirs(args.output_dir, exist_ok=True)
    t1 = os.path.join(args.output_dir, f"关键词调研_{args.date}.xlsx")
    t2 = os.path.join(args.output_dir, f"反查流量词_{args.date}.xlsx")
    t3 = os.path.join(args.output_dir, f"查广告词_{args.date}.xlsx")
    write_table1(t1)
    write_table2(t2)
    write_table3(t3)
    print(f"✅ 已生成测试原始表：\n  {t1}\n  {t2}\n  {t3}")


if __name__ == "__main__":
    main()
